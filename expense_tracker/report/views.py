import datetime
from io import BytesIO

from django.shortcuts import render
from django.http import HttpResponse
from django.views import View
from django.db.models import Sum
from django.utils import timezone
from django.core.exceptions import ValidationError

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.contrib.auth.decorators import login_required
# Import the models you want to report on
from expense.models import MoneyAllocation, Transaction, LoanRequest, ExpenseCategory
from account.models import User
from django.contrib.auth.decorators import user_passes_test
from django.utils.decorators import method_decorator
from django.views import View
from django.shortcuts import render
from django.http import HttpResponseForbidden
def is_admin(user):
    return hasattr(user, 'user_type') and user.user_type == 'admin'


@method_decorator(user_passes_test(is_admin, login_url='/users/login/', redirect_field_name=None), name='dispatch')
class FinanceReportView(View):
    def get(self, request):
        # Prepare a list of user types for the drop–down.
        # (Assuming your User model defines a CHOICES_USER_TYPE tuple.)
        user_types = [choice[0] for choice in User.CHOICES_USER_TYPE]
        return render(request, 'report.html', {'user_types': user_types})

    def post(self, request):
        # Retrieve filter values from the form submission
        user_type = request.POST.get('user_type')  # may be empty
        model_choice = request.POST.get('model_choice')
        date_range_option = request.POST.get('date_range_option')
        custom_start_date = request.POST.get('custom_start_date')
        custom_end_date = request.POST.get('custom_end_date')

        now = timezone.now()

        # Determine the date range based on the option selected.
        if date_range_option == 'today':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = now
        elif date_range_option == 'weekly':
            start_date = now - datetime.timedelta(days=7)
            end_date = now
        elif date_range_option == 'monthly':
            start_date = now - datetime.timedelta(days=30)
            end_date = now
        elif date_range_option == 'yearly':
            start_date = now - datetime.timedelta(days=365)
            end_date = now
        elif date_range_option == 'custom' and custom_start_date and custom_end_date:
            try:
                start_date = datetime.datetime.strptime(custom_start_date, '%Y-%m-%d')
                end_date = datetime.datetime.strptime(custom_end_date, '%Y-%m-%d')
            except ValueError:
                raise ValidationError("Custom dates must be in YYYY-MM-DD format.")
        else:
            start_date = end_date = None

        # Prepare the queryset based on the selected model and filters.
        data = []
        report_title = ''

        if model_choice == 'moneyallocation':
            qs = MoneyAllocation.objects.all()
            if start_date and end_date:
                qs = qs.filter(created_at__range=(start_date, end_date))
            if user_type:
                qs = qs.filter(allocated_by__user_type=user_type)
            data = qs
            report_title = 'Money Allocation'
        elif model_choice == 'transaction':
            qs = Transaction.objects.all()
            if start_date and end_date:
                qs = qs.filter(created_at__range=(start_date, end_date))
            if user_type:
                qs = qs.filter(user__user_type=user_type)
            data = qs
            report_title = 'Transaction'
        elif model_choice == 'loanrequest':
            qs = LoanRequest.objects.all()
            if start_date and end_date:
                qs = qs.filter(requested_at__range=(start_date, end_date))
            # For example, filter by the department (from_department) if a user type is provided.
            if user_type:
                qs = qs.filter(from_department=user_type)
            data = qs
            report_title = 'Loan Request'
        elif model_choice == 'expensecategory':
            qs = ExpenseCategory.objects.all()
            if start_date and end_date:
                qs = qs.filter(created_at__range=(start_date, end_date))
            if user_type:
                qs = qs.filter(created_by__user_type=user_type)
            data = qs
            report_title = 'Expense Category'
        else:
            # If no valid model is selected, you might want to return an error.
            return HttpResponse("Invalid model selection.", status=400)

        # Build the Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Finance Report"

        # === Custom Header ===
        # Merge cells A1:E1 and set a header title.
        header_title = "Finance Report from TED & S2L"
        ws.merge_cells('A1:E1')
        header_cell = ws['A1']
        header_cell.value = header_title
        header_cell.font = Font(size=16, bold=True)
        header_cell.alignment = Alignment(horizontal='center')

        # Write filter summary (row 2)
        ws.merge_cells('A2:E2')
        filter_text = f"Model: {report_title} | Date Range: {date_range_option.capitalize()}"
        if date_range_option == 'custom' and custom_start_date and custom_end_date:
            filter_text += f" ({custom_start_date} to {custom_end_date})"
        if user_type:
            filter_text += f" | User Type: {user_type}"
        ws['A2'].value = filter_text
        ws['A2'].font = Font(italic=True)
        ws['A2'].alignment = Alignment(horizontal='center')

        # === Data Table Headers ===
        start_row = 4
        headers = []
        if model_choice == 'moneyallocation':
            headers = ['ID', 'Allocated By', 'Allocated To', 'Amount', 'Source', 'Created At']
        elif model_choice == 'transaction':
            headers = ['ID', 'User', 'Amount', 'Category', 'Description', 'Source', 'Created At']
        elif model_choice == 'loanrequest':
            headers = ['ID', 'From Department', 'To Department', 'Amount', 'Status', 'Requested At', 'Approved At', 'Approved By']
        elif model_choice == 'expensecategory':
            headers = ['ID', 'Created By', 'Name', 'Description', 'Created At']

        # Write header row with basic styling.
        for col_num, column_title in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="CCCCCC")
            cell.alignment = Alignment(horizontal='center')

        # === Data Rows ===
        current_row = start_row + 1
        for item in data:
            col_num = 1
            if model_choice == 'moneyallocation':
                ws.cell(row=current_row, column=col_num, value=item.id)
                col_num += 1
                ws.cell(row=current_row, column=col_num, value=str(item.allocated_by))
                col_num += 1
                ws.cell(row=current_row, column=col_num, value=str(item.allocated_to))
                col_num += 1
                ws.cell(row=current_row, column=col_num, value=float(item.amount))
                col_num += 1
                ws.cell(row=current_row, column=col_num, value=item.source)
                col_num += 1
                ws.cell(row=current_row, column=col_num, value=item.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            elif model_choice == 'transaction':
                ws.cell(row=current_row, column=col_num, value=item.id)
                col_num += 1
                ws.cell(row=current_row, column=col_num, value=str(item.user))
                col_num += 1
                ws.cell(row=current_row, column=col_num, value=float(item.ammount))
                col_num += 1
                ws.cell(row=current_row, column=col_num, value=str(item.category) if item.category else '')
                col_num += 1
                ws.cell(row=current_row, column=col_num, value=str(item.category.description) if item.category else '')
                col_num += 1
                ws.cell(row=current_row, column=col_num, value=item.source)
                col_num += 1
                ws.cell(row=current_row, column=col_num, value=item.created_at.strftime('%Y-%m-%d %H:%M:%S'))

            elif model_choice == 'loanrequest':
                ws.cell(row=current_row, column=col_num, value=item.id)
                col_num += 1
                ws.cell(row=current_row, column=col_num, value=item.get_from_department_display())
                col_num += 1
                ws.cell(row=current_row, column=col_num, value=item.get_to_department_display())
                col_num += 1
                ws.cell(row=current_row, column=col_num, value=float(item.amount))
                col_num += 1

                # Write status with custom color styling.
                status_cell = ws.cell(row=current_row, column=col_num, value=item.status.capitalize())
                if item.status == 'pending':
                    status_cell.fill = PatternFill("solid", fgColor="FFFF00")  # Yellow for pending
                elif item.status == 'approved':
                    status_cell.fill = PatternFill("solid", fgColor="00FF00")  # Green for approved
                elif item.status == 'declined':
                    status_cell.fill = PatternFill("solid", fgColor="FF0000")  # Red for declined
                col_num += 1

                ws.cell(row=current_row, column=col_num, value=item.requested_at.strftime('%Y-%m-%d %H:%M:%S'))
                col_num += 1
                ws.cell(row=current_row, column=col_num, value=item.approved_at.strftime('%Y-%m-%d %H:%M:%S') if item.approved_at else '')
                col_num += 1
                ws.cell(row=current_row, column=col_num, value=str(item.approved_by) if item.approved_by else '')
            elif model_choice == 'expensecategory':
                ws.cell(row=current_row, column=col_num, value=item.id)
                col_num += 1
                ws.cell(row=current_row, column=col_num, value=str(item.created_by))
                col_num += 1
                ws.cell(row=current_row, column=col_num, value=item.name)
                col_num += 1
                ws.cell(row=current_row, column=col_num, value=item.description)
                col_num += 1
                ws.cell(row=current_row, column=col_num, value=item.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            current_row += 1

        # === Aggregations ===
        agg_start_row = current_row + 1
        ws.cell(row=agg_start_row, column=1, value="Aggregations:")
        ws.cell(row=agg_start_row, column=1).font = Font(bold=True)

        if model_choice == 'moneyallocation':
            total_amount = data.aggregate(total=Sum('amount'))['total'] or 0
            ws.cell(row=agg_start_row+1, column=1, value="Total Allocated Amount:")
            ws.cell(row=agg_start_row+1, column=2, value=float(total_amount))
        elif model_choice == 'transaction':
            total_amount = data.aggregate(total=Sum('ammount'))['total'] or 0
            ws.cell(row=agg_start_row+1, column=1, value="Total Transaction Amount:")
            ws.cell(row=agg_start_row+1, column=2, value=float(total_amount))
        elif model_choice == 'loanrequest':
            total_amount = data.aggregate(total=Sum('amount'))['total'] or 0
            pending_count = data.filter(status='pending').count()
            approved_count = data.filter(status='approved').count()
            declined_count = data.filter(status='declined').count()
            ws.cell(row=agg_start_row+1, column=1, value="Total Loan Amount:")
            ws.cell(row=agg_start_row+1, column=2, value=float(total_amount))
            ws.cell(row=agg_start_row+2, column=1, value="Pending Loans:")
            ws.cell(row=agg_start_row+2, column=2, value=pending_count)
            ws.cell(row=agg_start_row+3, column=1, value="Approved Loans:")
            ws.cell(row=agg_start_row+3, column=2, value=approved_count)
            ws.cell(row=agg_start_row+4, column=1, value="Declined Loans:")
            ws.cell(row=agg_start_row+4, column=2, value=declined_count)
        elif model_choice == 'expensecategory':
            count = data.count()
            ws.cell(row=agg_start_row+1, column=1, value="Total Expense Categories:")
            ws.cell(row=agg_start_row+1, column=2, value=count)

        # === Auto–Adjust Column Widths ===
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except Exception:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save the workbook to an in–memory bytes buffer and return it as an Excel file
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f'attachment; filename={report_title}_finance_report.xlsx'
        return response




from django.shortcuts import render
from django.db.models import Sum
from django.utils import timezone
@login_required
def dashboard(request):
    user = request.user  # Current user
    balance = user.balance  # User balance

    # Total spending (sum of all transaction amounts for the user)
    total_spending_data = Transaction.objects.filter(user=user).aggregate(total=Sum('ammount'))
    total_spending = total_spending_data.get('total') or 0

    # Today's expense (sum of all transactions recorded today)
    today = timezone.now().date()
    today_expense_data = Transaction.objects.filter(
        user=user, created_at__date=today
    ).aggregate(total=Sum('ammount'))
    today_expense = today_expense_data.get('total') or 0

    # Allocated money to the user today (excluding self-allocations)
    allocation_data = MoneyAllocation.objects.filter(
        allocated_to=user,
        created_at__date=today
    ).exclude(allocated_by=user).aggregate(total_allocated=Sum('amount'))
    total_allocated_money = allocation_data.get('total_allocated') or 0

    # Initialize total_expense
    total_expense = 0

    if user.user_type == 'admin':
        # 1. Get the total allocated money across all allocations.
        total_allocated_data = MoneyAllocation.objects.aggregate(
            total_allocated=Sum('amount')
        )
        total_allocated_money = total_allocated_data.get('total_allocated') or 0

        # 2. Get the total allocated money for allocations where the receiving user has user_type 'admin'
        admin_allocated_data = MoneyAllocation.objects.filter(
            allocated_to__user_type='admin'
        ).aggregate(total_admin_allocated=Sum('amount'))
        total_admin_allocated_money = admin_allocated_data.get('total_admin_allocated') or 0

        # Calculate total expense for admin
        total_expense = total_spending + total_allocated_money - total_admin_allocated_money
        total_allocated_money = total_admin_allocated_money

    else:
        # For non-admin users, total expense is total spending
        total_expense = total_spending
   # Count today's transactions created by each user_type ('s2l', 'ted', 'admin')
    today_s2l_transactions = Transaction.objects.filter(
        user__user_type='s2l', created_at__date=today
    ).count()
    today_ted_transactions = Transaction.objects.filter(
        user__user_type='ted', created_at__date=today
    ).count()
    today_admin_transactions = Transaction.objects.filter(
        user__user_type='admin', created_at__date=today
    ).count()
    total_admin_transactions = Transaction.objects.filter(user__user_type='admin').count()
        # Get recently created 10 transactions
    recent_transactions = Transaction.objects.filter(user=user).order_by('-created_at')[:10]

    # Get recently created 5 money allocations
    recent_allocations = MoneyAllocation.objects.filter(allocated_to=user).order_by('-created_at')[:5]
    # Count today's transactions created by each user_type ('s2l', 'ted')
    today_s2l_transactions = Transaction.objects.filter(
        user__user_type='s2l', created_at__date=today
    ).count()
    today_ted_transactions = Transaction.objects.filter(
        user__user_type='ted', created_at__date=today
    ).count()

    # Prepare data for the pie chart
    transactions_labels = ['S2L', 'TED']
    transactions_data = [today_s2l_transactions, today_ted_transactions]
    
    show_row =  user.user_type == 'admin'
    context = {
        'balance': balance,
        'total_expense': total_expense,
        'today_expense': today_expense,
        'today_allocated_money': total_allocated_money,
        'total_admin_transactions': total_admin_transactions,
        'today_s2l_transactions': today_s2l_transactions,
        'today_ted_transactions': today_ted_transactions,
        'today_admin_transactions': today_admin_transactions,
        'recent_transactions': recent_transactions,
        'recent_allocations': recent_allocations,
        'transactions_labels': transactions_labels,
        'transactions_data': transactions_data,
        "show_row": show_row,
    }

    return render(request, 'admin_dashboard.html', context)
